#!/usr/bin/env python3
"""生成颗粒课导入示例模板（根目录）。"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header(cell):
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )


def style_note(cell, lines):
    cell.font = Font(size=10, color="808080", italic=True)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def style_data(cell):
    cell.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    wb = Workbook()

    # 课程基本信息
    ws = wb.active
    ws.title = "课程基本信息"
    headers = ["课程名称 *", "适用专业", "课程简介", "难度", "预计课时", "学习目标", "关联知识点", "课程资源", "所属批次"]
    widths = [28, 24, 42, 10, 12, 42, 28, 28, 20]

    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "适用专业：从「专业字典」Sheet 选取，匹配则关联，不匹配则忽略\n"
        "难度：1-5，1 最易，5 最难\n"
        "预计课时：数字，单位小时\n"
        "关联知识点：从「知识点库」Sheet 选取，多个逗号分隔；匹配则关联，不匹配则自动新建并关联\n"
        "课程资源：从「任务资源库」Sheet 选取，多个逗号分隔；匹配则关联，不匹配则自动新建为文档类型资源并关联\n"
        "所属批次：从「批次字典」Sheet 选取，匹配则关联，不匹配则忽略\n"
        "导入后默认状态为 draft"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    style_note(ws.cell(row=1, column=1), note)
    ws.cell(row=1, column=1).value = note
    ws.row_dimensions[1].height = 130

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    sample_data = [
        [
            "Python 基础语法",
            "软件工程",
            "面向初学者的 Python 语法入门课程",
            2,
            8,
            "掌握 Python 基本语法、变量、控制流和函数定义",
            "Python 语法, 变量与数据类型, 流程控制",
            "Python 官方文档, Python 入门视频教程",
            "2024 春季软件工程批次",
        ],
        [
            "Web 安全基础",
            "网络工程",
            "Web 应用常见安全漏洞与防御基础",
            3,
            12,
            "理解 OWASP Top 10，掌握常见 Web 漏洞原理与防御方法",
            "OWASP, SQL 注入, XSS 攻击",
            "Web 安全实战手册, DVWA 实验环境",
            "2024 春季网络安全批次",
        ],
    ]
    for ridx, row in enumerate(sample_data, 3):
        for cidx, val in enumerate(row, 1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            style_data(cell)
        ws.row_dimensions[ridx].height = 40

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # 参考表数据
    ref_sheets = [
        (
            "【参考】专业字典",
            ["专业名称", "专业编码"],
            [32, 18],
            "仅作参考，无需编辑修改。\n课程基本信息 Sheet「适用专业」与本表名称一致则关联已有，不一致则忽略（不新建专业）。",
            [
                ["软件工程", "SE"],
                ["网络工程", "NE"],
                ["计算机科学与技术", "CS"],
                ["信息安全", "IS"],
                ["数据科学与大数据技术", "DS"],
            ],
        ),
        (
            "【参考】批次字典",
            ["批次名称"],
            [36],
            "仅作参考，无需编辑修改。\n课程基本信息 Sheet「所属批次」与本表名称一致则关联已有，不一致则忽略（不新建批次）。",
            [
                ["2024 春季软件工程批次"],
                ["2024 春季网络安全批次"],
                ["2024 秋季人工智能批次"],
            ],
        ),
        (
            "【参考】知识点库",
            ["知识点名称"],
            [36],
            "仅作参考，无需编辑修改。\n课程基本信息 Sheet「关联知识点」与本表名称一致则关联已有，不一致则自动新建并关联。",
            [
                ["Python 语法"],
                ["变量与数据类型"],
                ["流程控制"],
                ["函数与模块"],
                ["OWASP"],
                ["SQL 注入"],
                ["XSS 攻击"],
                ["CSRF 防护"],
            ],
        ),
        (
            "【参考】任务资源库",
            ["资源名称", "资源类型"],
            [42, 16],
            "仅作参考，无需编辑修改。\n课程基本信息 Sheet「课程资源」与本表一致则关联已有，不一致则自动新建为文档类型资源并关联。",
            [
                ["Python 官方文档", "文档"],
                ["Python 入门视频教程", "视频"],
                ["Web 安全实战手册", "文档"],
                ["DVWA 实验环境", "链接"],
            ],
        ),
    ]

    for title, cols, widths, note_text, rows in ref_sheets:
        rws = wb.create_sheet(title=title)
        rws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        style_note(rws.cell(row=1, column=1), note_text)
        rws.cell(row=1, column=1).value = note_text
        rws.row_dimensions[1].height = 60

        for col, h in enumerate(cols, 1):
            cell = rws.cell(row=2, column=col, value=h)
            style_header(cell)
            rws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

        for ridx, row in enumerate(rows, 3):
            for cidx, val in enumerate(row, 1):
                cell = rws.cell(row=ridx, column=cidx, value=val)
                style_data(cell)
            rws.row_dimensions[ridx].height = 24

        rws.freeze_panes = "A3"
        rws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

    wb.save("颗粒课批量导入模板（示例数据）.xlsx")


if __name__ == "__main__":
    main()
