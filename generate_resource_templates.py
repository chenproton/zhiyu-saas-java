#!/usr/bin/env python3
"""Generate sample Excel import templates for portal resources."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def add_note(sheet, cols, text):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = sheet.cell(row=1, column=1, value=text)
    cell.font = Font(size=10, color="808080", italic=True)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.row_dimensions[1].height = max(36, (text.count("\n") + 2) * 14)


def add_headers(sheet, headers, widths):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=2, column=i, value=h)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = align
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.row_dimensions[2].height = 28
    sheet.freeze_panes = "A3"


def add_row(sheet, row_idx, values):
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    align = Alignment(vertical="top", wrap_text=True)
    for i, v in enumerate(values, start=1):
        cell = sheet.cell(row=row_idx, column=i, value=v)
        cell.border = border
        cell.alignment = align
    sheet.row_dimensions[row_idx].height = 24


def create_industries():
    wb = Workbook()
    ws = wb.active
    ws.title = "行业列表"
    headers = ["行业代码 *", "行业名称 *", "上级行业代码", "排序", "是否启用"]
    widths = [18, 28, 18, 10, 12]
    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "行业代码：租户内唯一，已存在则更新。\n"
        "上级行业代码：填写本 Sheet 中已有的行业代码，或系统中已存在的行业代码；为空表示顶级行业。\n"
        "排序：数字，越小越靠前，默认为 0。\n"
        "是否启用：true/是/启用 表示启用；false/否/禁用 表示禁用，默认启用。"
    )
    add_note(ws, 5, note)
    add_headers(ws, headers, widths)

    data = [
        ("IT", "信息技术", "", 1, "是"),
        ("SW", "软件开发", "IT", 1, "是"),
        ("AI", "人工智能", "IT", 2, "是"),
        ("NET", "网络工程", "IT", 3, "是"),
        ("MFG", "智能制造", "", 2, "是"),
        ("ROB", "工业机器人", "MFG", 1, "是"),
        ("NEV", "新能源汽车", "", 3, "是"),
        ("BMS", "电池管理系统", "NEV", 1, "是"),
        ("MOTOR", "电驱动系统", "NEV", 2, "是"),
        ("EDU", "教育培训", "", 4, "是"),
    ]
    for i, row in enumerate(data, start=3):
        add_row(ws, i, row)

    wb.save("industry_import_demo.xlsx")
    print("Created industry_import_demo.xlsx")


def create_majors():
    wb = Workbook()
    ws = wb.active
    ws.title = "专业列表"
    headers = ["专业代码 *", "专业名称 *", "别名", "所属组织节点", "是否启用"]
    widths = [18, 30, 24, 28, 12]
    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "专业代码：租户内唯一，已存在则更新。\n"
        "所属组织节点：填写系统中已存在的组织节点名称，匹配则关联，不匹配则忽略。\n"
        "是否启用：true/是/启用 表示启用；false/否/禁用 表示禁用，默认启用。"
    )
    add_note(ws, 5, note)
    add_headers(ws, headers, widths)

    data = [
        ("CS101", "计算机科学与技术", "计科", "信息工程学院", "是"),
        ("SE202", "软件工程", "软工", "信息工程学院", "是"),
        ("AI303", "人工智能技术应用", "AI", "信息工程学院", "是"),
        ("NE401", "新能源汽车技术", "新能源", "汽车工程学院", "是"),
        ("BMS501", "动力电池技术", "电池", "汽车工程学院", "是"),
        ("ROB601", "工业机器人技术", "机器人", "智能制造学院", "是"),
        ("ME701", "机械设计与制造", "机制", "智能制造学院", "是"),
        ("BM801", "工商企业管理", "工管", "经济管理学院", "是"),
        ("AC901", "大数据与会计", "会计", "经济管理学院", "是"),
    ]
    for i, row in enumerate(data, start=3):
        add_row(ws, i, row)

    wb.save("major_import_demo.xlsx")
    print("Created major_import_demo.xlsx")


def create_organizations():
    wb = Workbook()
    ws = wb.active
    ws.title = "组织架构"
    headers = ["组织名称 *", "组织类型 *", "父组织名称", "排序"]
    widths = [30, 20, 30, 10]
    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "组织类型：须与系统中已存在的组织类型名称完全一致。\n"
        "父组织名称：填写本 Sheet 中已有的组织名称，或系统中已存在的组织名称；为空表示一级节点。\n"
        "排序：数字，越小越靠前，默认为 0。\n"
        "相同「组织名称+组织类型」的组合视为同一组织，重复导入会更新父组织和排序。"
    )
    add_note(ws, 4, note)
    add_headers(ws, headers, widths)

    data = [
        ("某某职业技术学院", "学校", "", 1),
        ("信息工程学院", "二级学院", "某某职业技术学院", 1),
        ("软件技术教研室", "行政职能部门", "信息工程学院", 1),
        ("软件技术2401班", "班级", "信息工程学院", 1),
        ("软件技术2402班", "班级", "信息工程学院", 2),
        ("人工智能2401班", "班级", "信息工程学院", 3),
        ("汽车工程学院", "二级学院", "某某职业技术学院", 2),
        ("新能源汽车2401班", "班级", "汽车工程学院", 1),
        ("新能源汽车2402班", "班级", "汽车工程学院", 2),
        ("智能制造学院", "二级学院", "某某职业技术学院", 3),
        ("工业机器人2401班", "班级", "智能制造学院", 1),
        ("经济管理学院", "二级学院", "某某职业技术学院", 4),
        ("大数据与会计2401班", "班级", "经济管理学院", 1),
    ]
    for i, row in enumerate(data, start=3):
        add_row(ws, i, row)

    wb.save("organization_import_demo.xlsx")
    print("Created organization_import_demo.xlsx")


def create_students():
    wb = Workbook()
    ws = wb.active
    ws.title = "学生列表"
    headers = ["登录账号(学号) *", "姓名 *", "密码 *", "班级(组织节点名称) *", "专业代码", "状态"]
    widths = [24, 16, 20, 28, 16, 12]
    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "登录账号(学号)：租户内唯一，已存在则跳过。\n"
        "密码：长度至少 8 位，且需同时包含字母和数字。\n"
        "班级：须与系统中已存在的组织节点名称完全一致。\n"
        "专业代码：填写系统中已存在的专业代码，匹配则关联，不匹配则忽略。\n"
        "状态：在籍 / 休学 / 退学 / 毕业 / 结业，默认为在籍。"
    )
    add_note(ws, 6, note)
    add_headers(ws, headers, widths)

    data = [
        ("S2024001", "张三", "Zhangsan2024", "软件技术2401班", "SE202", "在籍"),
        ("S2024002", "李四", "Lisi2024abc", "软件技术2401班", "SE202", "在籍"),
        ("S2024003", "王五", "Wangwu2024", "软件技术2402班", "SE202", "在籍"),
        ("S2024004", "赵六", "Zhaoliu2024", "人工智能2401班", "AI303", "在籍"),
        ("S2024005", "孙七", "Sunqi2024", "新能源汽车2401班", "NE401", "在籍"),
        ("S2024006", "周八", "Zhouba2024", "新能源汽车2402班", "NE401", "在籍"),
        ("S2024007", "吴九", "Wujiu2024", "工业机器人2401班", "ROB601", "在籍"),
        ("S2024008", "郑十", "Zhengshi2024", "大数据与会计2401班", "AC901", "在籍"),
        ("S2024009", "钱十一", "Qian1112024", "软件技术2401班", "SE202", "在籍"),
        ("S2024010", "冯十二", "Feng12_2024", "人工智能2401班", "AI303", "在籍"),
    ]
    for i, row in enumerate(data, start=3):
        add_row(ws, i, row)

    wb.save("student_import_demo.xlsx")
    print("Created student_import_demo.xlsx")


def create_teachers():
    wb = Workbook()
    ws = wb.active
    ws.title = "教师列表"
    headers = ["登录账号(工号) *", "姓名 *", "密码 *", "所属组织节点(名称)", "职位(逗号分隔)", "状态"]
    widths = [24, 16, 20, 28, 28, 12]
    note = (
        "填写说明：\n"
        "* 必填列。\n"
        "登录账号(工号)：租户内唯一，已存在则跳过。\n"
        "密码：长度至少 8 位，且需同时包含字母和数字。\n"
        "所属组织节点：填写系统中已存在的组织节点名称，匹配则关联，不匹配则忽略。\n"
        "职位：多个职位用逗号分隔，须与系统中已存在的职位名称一致，不匹配则忽略。\n"
        "状态：在职 / 离职 / 外聘 / 禁用，默认为在职。"
    )
    add_note(ws, 6, note)
    add_headers(ws, headers, widths)

    data = [
        ("T001", "陈教授", "Chen2024prof", "信息工程学院", "教授,专业带头人", "在职"),
        ("T002", "刘老师", "Liu2024tea", "信息工程学院", "讲师", "在职"),
        ("T003", "王老师", "Wang2024tea", "汽车工程学院", "副教授", "在职"),
        ("T004", "赵老师", "Zhao2024tea", "智能制造学院", "讲师", "在职"),
        ("T005", "孙老师", "Sun2024tea", "经济管理学院", "助教", "在职"),
        ("T006", "周工", "Zhou2024eng", "汽车工程学院", "企业导师", "外聘"),
        ("T007", "吴老师", "Wu2024tea", "信息工程学院", "讲师", "在职"),
        ("T008", "郑老师", "Zheng2024tea", "智能制造学院", "教授", "在职"),
    ]
    for i, row in enumerate(data, start=3):
        add_row(ws, i, row)

    wb.save("teacher_import_demo.xlsx")
    print("Created teacher_import_demo.xlsx")


if __name__ == "__main__":
    create_industries()
    create_majors()
    create_organizations()
    create_students()
    create_teachers()
